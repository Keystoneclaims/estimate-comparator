import csv
import io
import re
from typing import Dict, Iterable, List
from openpyxl import load_workbook
from .schemas import EstimateLine

HEADER_ALIASES = {
    "room": ["room", "area", "location", "loc", "trade room"],
    "category": ["category", "cat", "trade", "division", "coverage"],
    "description": ["description", "desc", "line item", "line_item", "item", "scope", "activity"],
    "quantity": ["quantity", "qty", "quan", "qnty"],
    "unit": ["unit", "uom", "measure"],
    "unit_price": ["unit_price", "unit price", "unit cost", "price", "rate"],
    "total": ["total", "rcv", "amount", "line total", "replacement cost", "recoverable total"],
}

def _clean_header(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())

def _money_to_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    try:
        return float(text)
    except ValueError:
        return 0.0

def _map_headers(headers: Iterable[str]) -> Dict[str, str]:
    cleaned = {_clean_header(h): h for h in headers}
    mapping = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in cleaned:
                mapping[canonical] = cleaned[alias]
                break
    return mapping

def _line_from_row(row: Dict[str, object], mapping: Dict[str, str], source: str) -> EstimateLine | None:
    description = str(row.get(mapping.get("description", ""), "") or "").strip()
    if not description:
        return None

    qty = _money_to_float(row.get(mapping.get("quantity", "")))
    unit_price = _money_to_float(row.get(mapping.get("unit_price", "")))
    total = _money_to_float(row.get(mapping.get("total", "")))
    if total == 0 and qty and unit_price:
        total = qty * unit_price

    return EstimateLine(
        source=source,
        room=str(row.get(mapping.get("room", ""), "") or "").strip(),
        category=str(row.get(mapping.get("category", ""), "") or "").strip(),
        description=description,
        quantity=qty,
        unit=str(row.get(mapping.get("unit", ""), "") or "").strip(),
        unit_price=unit_price,
        total=total,
        raw={str(k): v for k, v in row.items()},
    )

def parse_csv_bytes(content: bytes, source: str) -> List[EstimateLine]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    mapping = _map_headers(reader.fieldnames)
    if "description" not in mapping:
        raise ValueError("Could not find a description column. Add a header like description, desc, line item, or item.")
    return [line for row in reader if (line := _line_from_row(row, mapping, source))]

def parse_xlsx_bytes(content: bytes, source: str) -> List[EstimateLine]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "") for c in rows[0]]
    mapping = _map_headers(headers)
    if "description" not in mapping:
        raise ValueError("Could not find a description column in the first row.")
    lines = []
    for raw_values in rows[1:]:
        row = {headers[i]: raw_values[i] if i < len(raw_values) else None for i in range(len(headers))}
        line = _line_from_row(row, mapping, source)
        if line:
            lines.append(line)
    return lines

def parse_estimate_file(filename: str, content: bytes, source: str) -> List[EstimateLine]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv_bytes(content, source)
    if lower.endswith(".xlsx"):
        return parse_xlsx_bytes(content, source)
    if lower.endswith(".pdf"):
        raise ValueError("PDF upload is planned, but not enabled in this starter. Export the estimate to CSV/XLSX for now.")
    raise ValueError("Unsupported file type. Use .csv or .xlsx for this starter version.")
