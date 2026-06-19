import csv
import io
import re
from typing import Dict, Iterable, List, Sequence

import pdfplumber
from openpyxl import load_workbook

from .schemas import EstimateLine

HEADER_ALIASES = {
    "room": ["room", "area", "location", "loc", "trade room"],
    "category": ["category", "cat", "trade", "division", "coverage"],
    "description": ["description", "desc", "line item", "line_item", "item", "scope", "activity", "activity desc"],
    "quantity": ["quantity", "qty", "quan", "qnty"],
    "unit": ["unit", "uom", "measure"],
    "unit_price": ["unit_price", "unit price", "unit cost", "price", "rate", "unit $", "unit cost"],
    "total": ["total", "rcv", "amount", "line total", "replacement cost", "recoverable total", "total rcv"],
}

COMMON_UNITS = {
    "EA", "SF", "LF", "SY", "SQ", "HR", "DAY", "MO", "CY", "CF", "YD", "GL", "GAL", "LB", "TN", "TON", "BAG", "BOX", "ROLL", "SET", "PR", "PAIR", "LS", "JOB"
}

SUMMARY_WORDS = {
    "subtotal", "total", "totals", "tax", "overhead", "profit", "depreciation", "deductible", "net claim", "replacement cost", "actual cash value", "rcv", "acv"
}

ROOM_HINTS = [
    "kitchen", "kitchen/breakfast area", "powder room", "bathroom", "master bathroom", "bedroom", "master bedroom", "living room", "dining room", "family room", "hallway", "foyer", "entry/foyer", "entry", "laundry", "garage", "closet", "basement", "attic", "office", "den", "pantry", "exterior", "roof", "main level", "main floor", "second floor", "first floor", "mud room", "debris removal", "mitigation", "general"
]

NUMBER_TOKEN = r"\(?-?\d[\d,]*(?:\.\d+)?\)?"


def _clean_header(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def _money_to_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    try:
        return float(text)
    except ValueError:
        return 0.0


def _map_headers(headers: Iterable[str]) -> Dict[str, str]:
    cleaned = {_clean_header(h): h for h in headers}
    mapping = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in cleaned:
                mapping[canonical] = cleaned[alias]
                break
    return mapping


def _line_from_row(row: Dict[str, object], mapping: Dict[str, str], source: str) -> EstimateLine | None:
    description = str(row.get(mapping.get("description", ""), "") or "").strip()
    if not description:
        return None

    qty = _money_to_float(row.get(mapping.get("quantity", "")))
    unit_price = _money_to_float(row.get(mapping.get("unit_price", "")))
    total = _money_to_float(row.get(mapping.get("total", "")))
    if total == 0 and qty and unit_price:
        total = qty * unit_price

    return EstimateLine(
        source=source,
        room=str(row.get(mapping.get("room", ""), "") or "").strip(),
        category=str(row.get(mapping.get("category", ""), "") or "").strip(),
        description=description,
        quantity=qty,
        unit=str(row.get(mapping.get("unit", ""), "") or "").strip(),
        unit_price=unit_price,
        total=round(total, 2),
        raw={str(k): v for k, v in row.items()},
    )


def parse_csv_bytes(content: bytes, source: str) -> List[EstimateLine]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    mapping = _map_headers(reader.fieldnames)
    if "description" not in mapping:
        raise ValueError("Could not find a description column. Add a header like description, desc, line item, or item.")
    return [line for row in reader if (line := _line_from_row(row, mapping, source))]


def parse_xlsx_bytes(content: bytes, source: str) -> List[EstimateLine]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "") for c in rows[0]]
    mapping = _map_headers(headers)
    if "description" not in mapping:
        raise ValueError("Could not find a description column in the first row.")
    lines = []
    for raw_values in rows[1:]:
        row = {headers[i]: raw_values[i] if i < len(raw_values) else None for i in range(len(headers))}
        line = _line_from_row(row, mapping, source)
        if line:
            lines.append(line)
    return lines


def _looks_like_summary(description: str) -> bool:
    desc = _clean_header(description)
    if len(desc) < 3:
        return True
    return any(word == desc or desc.startswith(word + " ") for word in SUMMARY_WORDS)


def _guess_room_from_line(text: str, current_room: str) -> str:
    s = re.sub(r"\s+", " ", (text or "").strip())
    if not s:
        return current_room

    continued = re.match(r"^(?:CONTINUED\s*-\s*)(?P<room>.+)$", s, re.I)
    if continued:
        return continued.group("room").strip()

    # Xactimate room headers usually look like "Kitchen Height: 8'" or "Subroom: Closet (2) Height: 8'".
    height = re.match(r"^(?:Subroom:\s*)?(?P<room>[A-Za-z][A-Za-z0-9 /&()\-]+?)\s+Height:\s*", s)
    if height:
        room = height.group("room").strip()
        if len(room) <= 60:
            return room

    clean = _clean_header(s).strip(":")
    if not clean or len(clean) > 55:
        return current_room
    if clean in ROOM_HINTS:
        return s.strip().strip(":")
    if clean.endswith("room") or clean.endswith("bath") or clean.endswith("bathroom") or clean.endswith("foyer"):
        return s.strip().strip(":")
    if s.lower().startswith(("room:", "area:", "location:")):
        return s.split(":", 1)[1].strip()
    return current_room


def _parse_category_line(text: str) -> str | None:
    s = re.sub(r"_+", " ", (text or "").strip())
    s = re.sub(r"\s+", " ", s).strip(" -")
    if not s:
        return None
    if re.fullmatch(r"[A-Za-z/& ]{3,35}", s) and text.count("_") >= 4:
        return s.title()
    return None


def _parse_table_rows(table: Sequence[Sequence[object]], source: str) -> List[EstimateLine]:
    lines: List[EstimateLine] = []
    if not table:
        return lines
    cleaned_rows = [[str(cell or "").strip() for cell in row] for row in table if any(str(cell or "").strip() for cell in row)]
    if len(cleaned_rows) < 2:
        return lines

    for header_index in range(min(5, len(cleaned_rows) - 1)):
        headers = cleaned_rows[header_index]
        mapping = _map_headers(headers)
        if "description" not in mapping:
            continue
        for values in cleaned_rows[header_index + 1:]:
            row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
            line = _line_from_row(row, mapping, source)
            if line and not _looks_like_summary(line.description):
                lines.append(line)
        if lines:
            return lines
    return lines


# Generic Xactimate text-line parser. It handles both common layouts:
# Carrier style: DESCRIPTION QTY RESET REMOVE REPLACE TAX O&P TOTAL
# PA/company style: DESCRIPTION QUANTITY UNIT PRICE TAX O&P RCV DEPREC. ACV
XACTIMATE_LINE_RE = re.compile(
    rf"^\s*(?P<num>\d{{1,4}})\.\s+"
    rf"(?P<desc>.+?)\s+"
    rf"(?P<qty>-?\d[\d,]*(?:\.\d+)?)\s*(?P<unit>[A-Za-z]{{1,8}})\s+"
    rf"(?P<numbers>{NUMBER_TOKEN}(?:\s+{NUMBER_TOKEN}){{2,8}})\s*$"
)


def _is_noise_line(text: str) -> bool:
    s = _clean_header(text)
    if not s:
        return True
    noise_prefixes = (
        "description ", "totals:", "total:", "insured:", "property:", "claim number:", "policy number:",
        "date of loss:", "date inspected:", "price list:", "estimate:", "opening statement", "page:",
        "aaa roof technologies", "au erbach-storm", "auerbach-storm", "restoration/service/remodel"
    )
    return s.startswith(noise_prefixes)


def _numbers_from_tail(numbers: str) -> List[float]:
    return [_money_to_float(x) for x in re.findall(NUMBER_TOKEN, numbers or "")]


def _parse_xactimate_line(text: str, source: str, current_room: str, current_category: str, table_style: str) -> EstimateLine | None:
    line = re.sub(r"\s+", " ", text or "").strip()
    if not line or _is_noise_line(line):
        return None
    match = XACTIMATE_LINE_RE.match(line)
    if not match:
        return None

    desc = match.group("desc").strip(" -–—")
    desc = re.sub(r"^[A-Z]{2,5}\s+[A-Z0-9._-]+\s+", "", desc).strip()
    if _looks_like_summary(desc):
        return None

    qty = _money_to_float(match.group("qty"))
    unit = match.group("unit").upper()
    nums = _numbers_from_tail(match.group("numbers"))
    if len(nums) < 3:
        return None

    if table_style == "rcv_acv":
        # Layout: unit price, tax, O&P, RCV, depreciation, ACV.
        unit_price = nums[0]
        total = nums[-3] if len(nums) >= 6 else nums[-1]
    elif table_style == "reset_remove_replace":
        # Layout: reset, remove, replace, tax, O&P, total.
        # There is no single unit price column. Use the total/quantity as an approximate unit price for comparison.
        total = nums[-1]
        unit_price = round(total / qty, 2) if qty else 0.0
    elif len(nums) >= 6:
        # If the header was not captured, most 6-number layouts are RCV/ACV style.
        unit_price = nums[0]
        total = nums[-3]
    else:
        total = nums[-1]
        unit_price = round(total / qty, 2) if qty else 0.0

    if not desc or not total:
        return None

    return EstimateLine(
        source=source,
        room=current_room,
        category=current_category or "PDF import",
        description=desc,
        quantity=qty,
        unit=unit,
        unit_price=unit_price,
        total=round(total, 2),
        raw={"pdf_text_line": line, "table_style": table_style, "tail_numbers": nums},
    )


def _merge_short_continuation(previous: EstimateLine, continuation: str) -> EstimateLine:
    extra = re.sub(r"\s+", " ", continuation or "").strip()
    if not extra or len(extra) > 90:
        return previous
    if re.match(r"^\d+\.\s+", extra) or _is_noise_line(extra):
        return previous
    if any(token in _clean_header(extra) for token in ["height:", "walls", "ceiling", "floor", "opens into", "missing wall", "window ", "door ", "page:"]):
        return previous
    # Avoid appending long explanatory notes that are not part of the line item name.
    if previous.description.lower().endswith(extra.lower()):
        return previous
    previous.description = f"{previous.description} {extra}".strip()
    return previous


def parse_pdf_bytes(content: bytes, source: str) -> List[EstimateLine]:
    """Best-effort PDF estimate parser.

    Supports text-based Xactimate-style PDFs with either RCV/ACV columns or reset/remove/replace columns.
    Scanned/image-only PDFs still need OCR.
    """
    lines: List[EstimateLine] = []
    seen = set()
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            # First try actual table extraction.
            for table in page.extract_tables() or []:
                for parsed in _parse_table_rows(table, source):
                    key = (parsed.room, parsed.description, parsed.quantity, parsed.unit, parsed.total)
                    if key not in seen:
                        seen.add(key)
                        lines.append(parsed)

            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            current_room = ""
            current_category = "PDF import"
            table_style = "unknown"
            last_index: int | None = None
            can_append_to_last = False

            for raw_line in text.splitlines():
                clean_line = raw_line.strip()
                if not clean_line:
                    continue
                if "DESCRIPTION QUANTITY UNIT PRICE" in clean_line.upper():
                    table_style = "rcv_acv"
                    can_append_to_last = False
                    continue
                if "DESCRIPTION QTY RESET REMOVE" in clean_line.upper():
                    table_style = "reset_remove_replace"
                    can_append_to_last = False
                    continue
                category = _parse_category_line(clean_line)
                if category:
                    current_category = category
                    can_append_to_last = False
                    continue
                old_room = current_room
                current_room = _guess_room_from_line(clean_line, current_room)
                if current_room != old_room:
                    can_append_to_last = False

                parsed = _parse_xactimate_line(clean_line, source, current_room, current_category, table_style)
                if parsed:
                    key = (parsed.room, parsed.description, parsed.quantity, parsed.unit, parsed.total)
                    if key not in seen:
                        seen.add(key)
                        lines.append(parsed)
                        last_index = len(lines) - 1
                        can_append_to_last = True
                    continue

                # Append short Xactimate line continuations such as "unfaced batt", "reset", or "High grade".
                if can_append_to_last and last_index is not None and not re.match(r"^\d+\.\s+", clean_line):
                    before = lines[last_index].description
                    lines[last_index] = _merge_short_continuation(lines[last_index], clean_line)
                    # Only allow one immediate continuation line unless another parser result appears.
                    can_append_to_last = False
                else:
                    can_append_to_last = False

    if not lines:
        raise ValueError(
            "I could not read line items from this PDF. It may be a scanned/image PDF or a format that needs custom parsing. "
            "Try exporting the estimate as a text-based PDF, CSV, or XLSX."
        )
    return lines


def parse_estimate_file(filename: str, content: bytes, source: str) -> List[EstimateLine]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv_bytes(content, source)
    if lower.endswith(".xlsx"):
        return parse_xlsx_bytes(content, source)
    if lower.endswith(".pdf"):
        return parse_pdf_bytes(content, source)
    raise ValueError("Unsupported file type. Use .pdf, .csv, .xls, or .xlsx.")
